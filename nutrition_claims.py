#!/usr/bin/env python
"""Aggregate a state School Nutrition claim export into per-school daily meal
averages, and merge those averages into a MealsCount district template.

The claim export ("Claim Data" sheet) has one row per
site / program year / claim month / meal type, with the meals served that month
and the number of operating days. The daily average for a school year is the
weighted average across the months it reported:

    avg daily meals = sum(Total Meals Served) / sum(Number of Operating Days)

Usage:

  # 1. Aggregate the raw export into per-school, per-year averages
  python nutrition_claims.py aggregate SchoolNutritionDataExport_SY23-PY253.xlsx \
      -o school_meal_averages.xlsx

  # 2. Merge one year of those averages into a district template
  python nutrition_claims.py merge mealscount-schools.xlsx \
      --averages school_meal_averages.xlsx --year "2024 - 2025" \
      -o mealscount-schools-2024-2025.xlsx

Step 2 also accepts the raw export directly for --averages, in which case the
aggregation is done on the fly.
"""

import argparse
import os.path
import re
import sys

import pandas
from openpyxl import load_workbook

CLAIM_SHEET = "Claim Data"

# Columns we need out of the claim export.
COL_SPONSOR_ID = "Sponsor ID"
COL_SPONSOR_NAME = "Sponsor Name"
COL_SITE_NBR = "SiteNbr"
COL_SITE_NAME = "SiteName"
COL_PROGRAM_CODE = "Program Code"
COL_PROGRAM_YEAR = "Program Year"
COL_CLAIM_MONTH = "Claim Month"
COL_MEAL_TYPE = "Meal Type"
COL_MEALS = "Total Meals Served"
COL_DAYS = "Number of Operating/Serving Days"

REQUIRED_COLUMNS = [
    COL_SPONSOR_ID, COL_SPONSOR_NAME, COL_SITE_NBR, COL_SITE_NAME,
    COL_PROGRAM_CODE, COL_PROGRAM_YEAR, COL_CLAIM_MONTH, COL_MEAL_TYPE,
    COL_MEALS, COL_DAYS,
]

SCHOOL_YEAR_PROGRAM = "School Nutrition Programs"
SCHOOL_YEAR_RE = re.compile(r"^\s*(\d{4})\s*-\s*(\d{4})\s*$")
SUMMER_YEAR_RE = re.compile(r"^\s*Summer\s+(\d{4})\s*$", re.I)

# Columns written to the district template.
TEMPLATE_CODE_COLUMN = "school_code"
TEMPLATE_BREAKFAST_COLUMN = "daily_breakfast_served"
TEMPLATE_LUNCH_COLUMN = "daily_lunch_served"


def normalize_year(value):
    """Return a canonical 'YYYY - YYYY' fiscal year, or None if unrecognized.

    Summer sessions are folded into the school year they follow, so
    'Summer 2023' becomes '2022 - 2023'.
    """
    if value is None:
        return None
    text = str(value).strip()
    match = SCHOOL_YEAR_RE.match(text)
    if match:
        return "%s - %s" % (match.group(1), match.group(2))
    match = SUMMER_YEAR_RE.match(text)
    if match:
        end = int(match.group(1))
        return "%i - %i" % (end - 1, end)
    return None


def parse_year_argument(value):
    """Accept '2024 - 2025', '2024-2025', 'SY2425' or '2425' for --year."""
    text = str(value).strip()
    year = normalize_year(text)
    if year:
        return year
    match = re.match(r"^(?:SY)?\s*(\d{2})\s*-?\s*(\d{2})$", text, re.I)
    if match:
        return "20%s - 20%s" % (match.group(1), match.group(2))
    raise argparse.ArgumentTypeError(
        "could not read '%s' as a fiscal year (try '2024 - 2025')" % value
    )


def normalize_code(value):
    """Normalize a site number / school code so the two files can be matched.

    The export stores SiteNbr as a number while templates often carry the
    zero-padded string, so both sides are reduced to an unpadded digit string.
    """
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return None
    # Excel reads an all-digit code as a float often enough to be worth handling
    if re.match(r"^\d+\.0$", text):
        text = text[:-2]
    stripped = text.lstrip("0")
    return stripped or "0"


def read_claim_data(filename):
    """Read the claim export, returning the 'Claim Data' rows."""
    sheets = pandas.ExcelFile(filename).sheet_names
    sheet = CLAIM_SHEET if CLAIM_SHEET in sheets else sheets[0]
    df = pandas.read_excel(filename, sheet_name=sheet)
    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise SystemExit(
            "%s (sheet '%s') is missing expected columns: %s"
            % (filename, sheet, ", ".join(missing))
        )
    return df


def aggregate_claims(df, include_summer=False, decimals=1):
    """Collapse monthly claim rows into per-school, per-fiscal-year averages."""
    df = df.copy()
    df["fiscal_year"] = df[COL_PROGRAM_YEAR].map(normalize_year)

    unknown = df["fiscal_year"].isna().sum()
    if unknown:
        print("Skipping %i rows with an unrecognized Program Year" % unknown)
    df = df[df["fiscal_year"].notna()]

    if not include_summer:
        is_school_year = df[COL_PROGRAM_YEAR].astype(str).str.match(SCHOOL_YEAR_RE)
        df = df[is_school_year & (df[COL_PROGRAM_CODE] == SCHOOL_YEAR_PROGRAM)]

    df = df[df[COL_MEAL_TYPE].isin(["Breakfast", "Lunch"])]
    df = df[df[COL_DAYS] > 0]
    if df.empty:
        raise SystemExit("No breakfast or lunch claim rows found to aggregate")

    keys = ["fiscal_year", COL_SPONSOR_ID, COL_SPONSOR_NAME,
            COL_SITE_NBR, COL_SITE_NAME]
    grouped = df.groupby(keys + [COL_MEAL_TYPE], dropna=False).agg(
        meals=(COL_MEALS, "sum"),
        operating_days=(COL_DAYS, "sum"),
        months_reported=(COL_CLAIM_MONTH, "nunique"),
    ).reset_index()

    wide = grouped.pivot_table(
        index=keys, columns=COL_MEAL_TYPE,
        values=["meals", "operating_days", "months_reported"],
    )
    wide.columns = ["%s_%s" % (meal.lower(), stat) for stat, meal in wide.columns]
    wide = wide.reset_index()

    for meal in ("breakfast", "lunch"):
        for stat in ("meals", "operating_days", "months_reported"):
            column = "%s_%s" % (meal, stat)
            if column not in wide.columns:
                wide[column] = 0
            wide[column] = wide[column].fillna(0).astype(int)
        days = wide["%s_operating_days" % meal]
        wide["avg_daily_%s" % meal] = (
            wide["%s_meals" % meal].where(days > 0, 0) / days.where(days > 0, 1)
        ).astype(float).round(decimals)

    wide = wide.rename(columns={
        COL_SPONSOR_ID: "sponsor_id",
        COL_SPONSOR_NAME: "sponsor_name",
        COL_SITE_NBR: "school_code",
        COL_SITE_NAME: "school_name",
    })
    columns = [
        "fiscal_year", "sponsor_id", "sponsor_name", "school_code", "school_name",
        "avg_daily_breakfast", "avg_daily_lunch",
        "breakfast_meals", "breakfast_operating_days", "breakfast_months_reported",
        "lunch_meals", "lunch_operating_days", "lunch_months_reported",
    ]
    return wide[columns].sort_values(["fiscal_year", "sponsor_name", "school_name"])


def load_averages(filename, include_summer=False, decimals=1):
    """Load an averages file, aggregating it first if it is a raw claim export."""
    sheets = pandas.ExcelFile(filename).sheet_names if filename.endswith(
        (".xlsx", ".xlsm")) else None
    if sheets and CLAIM_SHEET in sheets:
        return aggregate_claims(
            read_claim_data(filename), include_summer=include_summer, decimals=decimals
        )
    if filename.endswith(".csv"):
        averages = pandas.read_csv(filename)
    else:
        averages = pandas.read_excel(filename)
    missing = [c for c in ("fiscal_year", "school_code", "avg_daily_breakfast",
                           "avg_daily_lunch") if c not in averages.columns]
    if missing:
        raise SystemExit(
            "%s is neither a claim export nor an averages file (missing %s)"
            % (filename, ", ".join(missing))
        )
    return averages


def write_table(df, filename, sheet_name="Averages"):
    if filename.endswith(".csv"):
        df.to_csv(filename, index=False)
    else:
        df.to_excel(filename, sheet_name=sheet_name, index=False)


def resolve_code_column(header, code_column, template, sheet_title):
    """Find the school code column, tolerating a template with a mangled header.

    Some exports of the district template come back with the first header cell
    blanked out or replaced by a stray character, so fall back to the first
    column rather than refusing to merge.
    """
    if code_column:
        if code_column not in header:
            raise SystemExit(
                "Template %s (sheet '%s') has no '%s' column; header is: %s"
                % (template, sheet_title, code_column,
                   ", ".join(str(h) for h in header))
            )
        return header.index(code_column) + 1
    if TEMPLATE_CODE_COLUMN in header:
        return header.index(TEMPLATE_CODE_COLUMN) + 1
    print("Template %s has no '%s' header (found %r), using the first column"
          % (template, TEMPLATE_CODE_COLUMN, header[0] if header else None))
    return 1


def merge_into_template(template, averages, year, output, decimals=0,
                        sponsor_id=None, dry_run=False, code_column=None):
    """Write the year's averages into a copy of the district template.

    Only the daily breakfast/lunch columns are touched; every other cell of the
    template (enrollment, eligibility, formatting) is carried through as-is.
    """
    averages = averages[averages["fiscal_year"] == year]
    if sponsor_id is not None:
        wanted = normalize_code(sponsor_id)
        averages = averages[averages["sponsor_id"].map(normalize_code) == wanted]
    if averages.empty:
        raise SystemExit("No aggregated rows for fiscal year %s" % year)

    by_code = {}
    for row in averages.to_dict("records"):
        code = normalize_code(row["school_code"])
        if code is not None:
            by_code[code] = row

    workbook = load_workbook(template)
    worksheet = workbook.active
    header = [cell.value for cell in worksheet[1]]
    for column in (TEMPLATE_BREAKFAST_COLUMN, TEMPLATE_LUNCH_COLUMN):
        if column not in header:
            raise SystemExit(
                "Template %s (sheet '%s') has no '%s' column"
                % (template, worksheet.title, column)
            )
    code_col = resolve_code_column(header, code_column, template, worksheet.title)
    breakfast_col = header.index(TEMPLATE_BREAKFAST_COLUMN) + 1
    lunch_col = header.index(TEMPLATE_LUNCH_COLUMN) + 1
    name_col = header.index("school_name") + 1 if "school_name" in header else None

    matched, unmatched = 0, []
    for row_number in range(2, worksheet.max_row + 1):
        raw_code = worksheet.cell(row=row_number, column=code_col).value
        code = normalize_code(raw_code)
        if code is None:
            continue
        row = by_code.get(code)
        if row is None:
            unmatched.append((
                raw_code,
                name_col and worksheet.cell(row=row_number, column=name_col).value,
            ))
            continue
        matched += 1
        breakfast = round(float(row["avg_daily_breakfast"]), decimals)
        lunch = round(float(row["avg_daily_lunch"]), decimals)
        if decimals <= 0:
            breakfast, lunch = int(breakfast), int(lunch)
        worksheet.cell(row=row_number, column=breakfast_col).value = breakfast
        worksheet.cell(row=row_number, column=lunch_col).value = lunch

    print("Merged %s into %i of %i template rows"
          % (year, matched, matched + len(unmatched)))
    if unmatched:
        print("No %s claim data for %i school(s), left unchanged:" % (year, len(unmatched)))
        for code, name in unmatched:
            print("  %s %s" % (code, name or ""))

    if dry_run:
        print("Dry run, %s not written" % output)
        return matched

    workbook.save(output)
    print("Wrote %s" % output)
    return matched


def command_aggregate(args):
    averages = aggregate_claims(
        read_claim_data(args.export),
        include_summer=args.include_summer,
        decimals=args.decimals,
    )
    write_table(averages, args.output)
    print("Wrote %s: %i school-year rows across %i schools and %i fiscal years"
          % (args.output, len(averages), averages["school_code"].nunique(),
             averages["fiscal_year"].nunique()))
    for year, count in averages["fiscal_year"].value_counts().sort_index().items():
        print("  %s: %i schools" % (year, count))
    return 0


def command_merge(args):
    # Aggregate at full precision here; the merge does the final rounding.
    averages = load_averages(args.averages, include_summer=args.include_summer,
                             decimals=6)
    merge_into_template(
        args.template, averages, args.year, args.output,
        decimals=args.decimals, sponsor_id=args.sponsor_id, dry_run=args.dry_run,
        code_column=args.code_column,
    )
    return 0


def main(argv=None):
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    subparsers = parser.add_subparsers(dest="command", required=True)

    aggregate = subparsers.add_parser(
        "aggregate", help="summarize a claim export into per-school daily averages")
    aggregate.add_argument("export", help="School Nutrition claim export (.xlsx)")
    aggregate.add_argument("-o", "--output", default="school_meal_averages.xlsx",
                           help="output .xlsx or .csv (default: %(default)s)")
    aggregate.add_argument("--decimals", type=int, default=1,
                           help="decimal places for the averages (default: %(default)s)")
    aggregate.add_argument("--include-summer", action="store_true",
                           help="also count summer program claims, folded into the "
                                "school year they follow")
    aggregate.set_defaults(func=command_aggregate)

    merge = subparsers.add_parser(
        "merge", help="write one year of averages into a district template")
    merge.add_argument("template", help="district template .xlsx to fill in")
    merge.add_argument("--averages", required=True,
                       help="averages file from `aggregate`, or a raw claim export")
    merge.add_argument("--year", required=True, type=parse_year_argument,
                       help="fiscal year to merge, e.g. '2024 - 2025'")
    merge.add_argument("-o", "--output", help="output .xlsx (default: <template>-<year>.xlsx)")
    merge.add_argument("--decimals", type=int, default=0,
                       help="decimal places for the merged counts (default: %(default)s)")
    merge.add_argument("--code-column",
                       help="template column holding the school code "
                            "(default: %s, falling back to the first column)"
                            % TEMPLATE_CODE_COLUMN)
    merge.add_argument("--sponsor-id",
                       help="only use claim rows for this sponsor/district id")
    merge.add_argument("--include-summer", action="store_true",
                       help="when aggregating a raw export, also count summer claims")
    merge.add_argument("--dry-run", action="store_true",
                       help="report what would change without writing the output")
    merge.set_defaults(func=command_merge)

    args = parser.parse_args(argv)
    if args.command == "merge" and not args.output:
        base = os.path.splitext(args.template)[0]
        args.output = "%s-%s.xlsx" % (base, args.year.replace(" ", ""))
    return args.func(args)


if __name__ == "__main__":
    sys.exit(main())
